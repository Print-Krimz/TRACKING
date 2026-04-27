"""
Report Service

Generates structured reports from application data.
Supports CSV, Excel (XLSX), and PDF output formats.

Report Types:
    - Pipeline: Application status breakdown per job
    - Match Scores: AI match score analytics per job
    - Usage: System activity metrics over time
"""

import csv
import io
import json
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional, Tuple

from sqlmodel import Session, select, func
from models.job import JobRequisition, JobStatus
from models.application import Application, ApplicationStatus
from models.resume import Resume
from models.user import User


# =============================================================================
# Report Data Generators
# =============================================================================

def generate_pipeline_report(
    session: Session,
    job_id: Optional[int] = None,
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
) -> Tuple[List[str], List[Dict[str, Any]]]:
    """
    Generate candidate pipeline report.

    Returns a breakdown of applications grouped by job and status.

    Args:
        session: Database session
        job_id: Optional filter to a single job
        date_from: Optional start date filter
        date_to: Optional end date filter

    Returns:
        Tuple of (column names, list of row dicts)
    """
    columns = [
        "Job Title", "Department", "Received", "Screening",
        "Interview", "Offer", "Hired", "Rejected", "Withdrawn", "Total"
    ]

    # Build job query
    job_query = select(JobRequisition)
    if job_id:
        job_query = job_query.where(JobRequisition.id == job_id)

    jobs = session.exec(job_query).all()

    rows = []
    for job in jobs:
        # Build application query for this job
        app_query = select(Application).where(Application.job_id == job.id)
        if date_from:
            app_query = app_query.where(Application.applied_at >= date_from)
        if date_to:
            app_query = app_query.where(Application.applied_at <= date_to)

        apps = session.exec(app_query).all()

        # Count by status
        status_counts = {}
        for status in ApplicationStatus:
            status_counts[status.value] = len(
                [a for a in apps if a.status == status]
            )

        rows.append({
            "Job Title": job.title,
            "Department": job.department or "N/A",
            "Received": status_counts.get("received", 0),
            "Screening": status_counts.get("screening", 0),
            "Interview": status_counts.get("interview", 0),
            "Offer": status_counts.get("offer", 0),
            "Hired": status_counts.get("hired", 0),
            "Rejected": status_counts.get("rejected", 0),
            "Withdrawn": status_counts.get("withdrawn", 0),
            "Total": len(apps),
        })

    return columns, rows


def generate_match_score_report(
    session: Session,
    job_id: Optional[int] = None,
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
) -> Tuple[List[str], List[Dict[str, Any]]]:
    """
    Generate AI match score analytics report.

    Returns aggregate match-score statistics per job, including
    the distribution of recommendation categories from score_breakdown.

    Args:
        session: Database session
        job_id: Optional filter to a single job
        date_from: Optional start date filter
        date_to: Optional end date filter

    Returns:
        Tuple of (column names, list of row dicts)
    """
    columns = [
        "Job Title", "Total Scored", "Avg Score", "Min Score", "Max Score",
        "Strong Match", "Good Match", "Partial Match", "Weak Match"
    ]

    job_query = select(JobRequisition)
    if job_id:
        job_query = job_query.where(JobRequisition.id == job_id)

    jobs = session.exec(job_query).all()

    rows = []
    for job in jobs:
        app_query = (
            select(Application)
            .where(Application.job_id == job.id)
            .where(Application.match_score != None)  # noqa: E711
        )
        if date_from:
            app_query = app_query.where(Application.applied_at >= date_from)
        if date_to:
            app_query = app_query.where(Application.applied_at <= date_to)

        scored_apps = session.exec(app_query).all()

        if not scored_apps:
            continue

        scores = [a.match_score for a in scored_apps]
        avg_score = round(sum(scores) / len(scores), 1)
        min_score = min(scores)
        max_score = max(scores)

        # Parse recommendation distribution from score_breakdown JSON
        rec_counts = {
            "strong_match": 0,
            "good_match": 0,
            "partial_match": 0,
            "weak_match": 0,
        }
        for app in scored_apps:
            if app.score_breakdown:
                try:
                    breakdown = json.loads(app.score_breakdown)
                    rec = breakdown.get("recommendation", "partial_match")
                    if rec in rec_counts:
                        rec_counts[rec] += 1
                except (json.JSONDecodeError, AttributeError):
                    pass

        rows.append({
            "Job Title": job.title,
            "Total Scored": len(scored_apps),
            "Avg Score": avg_score,
            "Min Score": min_score,
            "Max Score": max_score,
            "Strong Match": rec_counts["strong_match"],
            "Good Match": rec_counts["good_match"],
            "Partial Match": rec_counts["partial_match"],
            "Weak Match": rec_counts["weak_match"],
        })

    return columns, rows


def generate_usage_report(
    session: Session,
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
    **kwargs,
) -> Tuple[List[str], List[Dict[str, Any]]]:
    """
    Generate system usage analytics report.

    Returns daily counts of key activities: resumes submitted,
    applications received, resumes analyzed, and jobs created.

    Args:
        session: Database session
        date_from: Optional start date (defaults to 30 days ago)
        date_to: Optional end date (defaults to now)

    Returns:
        Tuple of (column names, list of row dicts)
    """
    columns = [
        "Date", "Resumes Submitted", "Applications Received",
        "Resumes Analyzed", "Jobs Created"
    ]

    if not date_from:
        date_from = datetime.utcnow() - timedelta(days=30)
    if not date_to:
        date_to = datetime.utcnow()

    # Fetch all relevant records in the date range
    resumes = session.exec(
        select(Resume).where(
            Resume.created_at >= date_from,
            Resume.created_at <= date_to,
        )
    ).all()

    applications = session.exec(
        select(Application).where(
            Application.applied_at >= date_from,
            Application.applied_at <= date_to,
        )
    ).all()

    jobs = session.exec(
        select(JobRequisition).where(
            JobRequisition.created_at >= date_from,
            JobRequisition.created_at <= date_to,
        )
    ).all()

    # Group by date
    rows = []
    current = date_from.replace(hour=0, minute=0, second=0, microsecond=0)
    end = date_to.replace(hour=23, minute=59, second=59, microsecond=999999)

    while current <= end:
        date_str = current.strftime("%Y-%m-%d")

        resumes_count = len([
            r for r in resumes
            if r.created_at.strftime("%Y-%m-%d") == date_str
        ])
        apps_count = len([
            a for a in applications
            if a.applied_at.strftime("%Y-%m-%d") == date_str
        ])
        analyzed_count = len([
            r for r in resumes
            if r.created_at.strftime("%Y-%m-%d") == date_str
            and r.analysis_result is not None
        ])
        jobs_count = len([
            j for j in jobs
            if j.created_at.strftime("%Y-%m-%d") == date_str
        ])

        rows.append({
            "Date": date_str,
            "Resumes Submitted": resumes_count,
            "Applications Received": apps_count,
            "Resumes Analyzed": analyzed_count,
            "Jobs Created": jobs_count,
        })

        current += timedelta(days=1)

    return columns, rows


# =============================================================================
# File Format Generators
# =============================================================================

def to_csv_bytes(columns: List[str], rows: List[Dict[str, Any]]) -> bytes:
    """
    Convert report data to CSV bytes.

    Args:
        columns: Column header names
        rows: List of row dictionaries

    Returns:
        CSV file content as bytes (UTF-8 with BOM for Excel compatibility)
    """
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=columns)
    writer.writeheader()
    for row in rows:
        writer.writerow(row)

    # UTF-8 BOM for proper Excel display
    return b'\xef\xbb\xbf' + output.getvalue().encode("utf-8")


def to_xlsx_bytes(
    columns: List[str],
    rows: List[Dict[str, Any]],
    title: str = "Report",
) -> bytes:
    """
    Convert report data to Excel XLSX bytes.

    Creates a styled workbook with a header row and auto-sized columns.

    Args:
        columns: Column header names
        rows: List of row dictionaries
        title: Worksheet title

    Returns:
        XLSX file content as bytes
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = title

    # Header style
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        bottom=Side(style="thin", color="C7D2FE")
    )

    # Write headers
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Write data rows
    data_font = Font(name="Calibri", size=10)
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))
            cell.font = data_font
            cell.border = thin_border

    # Auto-size columns
    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(str(col_name))
        for row_data in rows:
            val_len = len(str(row_data.get(col_name, "")))
            if val_len > max_len:
                max_len = val_len
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 4

    # Freeze header row
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def to_pdf_bytes(
    columns: List[str],
    rows: List[Dict[str, Any]],
    title: str = "Report",
) -> bytes:
    """
    Convert report data to PDF bytes.

    Creates a landscape PDF with a styled table.

    Args:
        columns: Column header names
        rows: List of row dictionaries
        title: Report title

    Returns:
        PDF file content as bytes
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=0.5 * inch,
        rightMargin=0.5 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
    )

    styles = getSampleStyleSheet()
    elements = []

    # Title
    title_style = styles["Title"]
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 12))

    # Subtitle with generation timestamp
    subtitle_style = styles["Normal"]
    elements.append(Paragraph(
        f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
        subtitle_style
    ))
    elements.append(Spacer(1, 20))

    # Build table data
    table_data = [columns]
    for row in rows:
        table_data.append([str(row.get(col, "")) for col in columns])

    if len(table_data) == 1:
        elements.append(Paragraph("No data available for the selected filters.", subtitle_style))
    else:
        # Calculate column widths to fit the page
        avail_width = landscape(A4)[0] - 1 * inch
        col_width = avail_width / len(columns)

        table = Table(table_data, colWidths=[col_width] * len(columns))
        table.setStyle(TableStyle([
            # Header
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            # Data
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ALIGN", (0, 1), (-1, -1), "CENTER"),
            # Grid
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E0E7FF")),
            # Alternate row colors
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [
                colors.white, colors.HexColor("#F5F3FF")
            ]),
            # Padding
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(table)

    doc.build(elements)
    return output.getvalue()


# =============================================================================
# Main Report Dispatcher
# =============================================================================

# Map of report type -> generator function
REPORT_GENERATORS = {
    "pipeline": generate_pipeline_report,
    "match_scores": generate_match_score_report,
    "usage": generate_usage_report,
}

# Map of format -> file converter
FORMAT_CONVERTERS = {
    "csv": to_csv_bytes,
    "xlsx": to_xlsx_bytes,
    "pdf": to_pdf_bytes,
}

# Report type metadata
REPORT_TYPES = [
    {
        "id": "pipeline",
        "name": "Candidate Pipeline Report",
        "description": "Breakdown of applications per job by their current status (Received, Screening, Interview, Offer, etc.)",
        "supported_formats": ["json", "csv", "xlsx", "pdf"],
        "supports_job_filter": True,
        "supports_date_filter": True,
    },
    {
        "id": "match_scores",
        "name": "AI Match Score Report",
        "description": "Aggregated AI match-score analytics per job including average scores and recommendation distribution.",
        "supported_formats": ["json", "csv", "xlsx", "pdf"],
        "supports_job_filter": True,
        "supports_date_filter": True,
    },
    {
        "id": "usage",
        "name": "System Usage Report",
        "description": "Daily system activity: resumes submitted, applications received, resumes analyzed, and jobs created.",
        "supported_formats": ["json", "csv", "xlsx", "pdf"],
        "supports_job_filter": False,
        "supports_date_filter": True,
    },
]


# Pretty titles for file names and PDF headers
REPORT_TITLES = {
    "pipeline": "Candidate Pipeline Report",
    "match_scores": "AI Match Score Report",
    "usage": "System Usage Report",
}
